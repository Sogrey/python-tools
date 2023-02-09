# 导入openpyxl库
import openpyxl

# 打开excel文件
wb = openpyxl.load_workbook('测试表模板.xlsx')

# 获取活动表
# sheet = wb.active
sheet = wb['绩效表']

# 输出表头
print('<table>')
print('<tr>')
for col in sheet.iter_cols(min_row=1, max_row=1):
    for cell in col:
        print('<th>{}</th>'.format(cell.value))
print('</tr>')

# 输出表格内容
for row in sheet.iter_rows(min_row=2):
    print('<tr>')
    for cell in row:
        # 判断单元格是否被合并，如果被合并，则取合并单元格的第一个单元格的值
        if cell.coordinate in sheet.merged_cells:
            cell = sheet.cell(sheet.merged_cells[cell.coordinate].min_row, sheet.merged_cells[cell.coordinate].min_col)
        print('<td>{}</td>'.format(cell.value))
    print('</tr>')

# 输出表尾
print('</table>')